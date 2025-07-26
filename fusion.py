
import os
import glob
import math
import shutil
import re
from datetime import datetime
from openpyxl import load_workbook
import time
import gc
import tempfile

log = print

def set_logger(custom_log):
    global log
    log = custom_log

def supprimer_et_attendre(chemin_fichier, timeout=30):
    if os.path.exists(chemin_fichier):
        # 🔄 Tente de libérer les ressources encore actives
        gc.collect()
        time.sleep(1)

        try:
            os.remove(chemin_fichier)
            print(f"🗑️ Fichier demandé pour suppression : {os.path.basename(chemin_fichier)}")
        except PermissionError as e:
            print(f"❌ Échec suppression immédiate : {e}")

    # 🔁 Attente active jusqu'à suppression réelle
    elapsed = 0
    while os.path.exists(chemin_fichier) and elapsed < timeout:
        time.sleep(1)
        elapsed += 1

    if not os.path.exists(chemin_fichier):
        print(f"✅ Suppression confirmée de : {os.path.basename(chemin_fichier)}")
    else:
        print(f"⚠️ Le fichier est encore présent après {timeout} secondes : {os.path.basename(chemin_fichier)}")

        
def get_x_fichiers_recents(dossier, pattern="*.xlsx", limit=2):
    fichiers = glob.glob(os.path.join(dossier, pattern))
    fichiers_trie = sorted(fichiers, key=os.path.getmtime, reverse=True)
    return fichiers_trie[:limit]

def arrondir_heure_par_15min(dt):
    heure = dt.hour
    minute = dt.minute
    quart = math.ceil(minute / 15) * 15
    if quart == 60:
        heure += 1
        quart = 0
    return f"{heure:02d}{quart:02d}"

def fusionner(choix, mois, annee):
    # ✅ Nouveau chemin persistant pour le dossier 'extract'
    dossier_extract = os.path.join(
        os.path.expanduser("~"),
        "Wavestone",
        "WO - CTO - CDM - Clôture",
        "extract"
    )
    print("🔍 Utilisateur courant :", os.getlogin())
    print("📂 Home path utilisé :", os.path.expanduser("~"))
    os.makedirs(dossier_extract, exist_ok=True)
    print(f"📁 Dossier extract utilisé : {dossier_extract}")

    
    
    # if choix == "Check imputations":

    log(f"🛠️ Réalisation de la mise à jour du fichier ({choix})...")
    
    if choix == "Suivi du TACE Timesheets":

        fichiers = glob.glob(os.path.join(dossier_extract, "*.xlsx"))
        fichiers_trie = sorted(fichiers, key=os.path.getctime, reverse=True)

        if len(fichiers_trie) < 2:
            print("❌ Moins de 2 fichiers trouvés pour la fusion.")
            return

        fichier_1, fichier_2 = fichiers_trie[:2]

        log(f"✅ Fusion de :\n- {os.path.basename(fichier_1)}\n- {os.path.basename(fichier_2)}")

        if "analyse des imputations" in os.path.basename(fichier_1).lower():
            fichier_dest = fichier_1
            fichier_source = fichier_2
        else:
            fichier_dest = fichier_2
            fichier_source = fichier_1

        # 🔄 Vider les colonnes A à G de la feuille "export_timesheet" à partir de la ligne 2
        wb_analyse = load_workbook(fichier_dest)
        ws_export = wb_analyse["export_timesheet"]

        for row in ws_export.iter_rows(min_row=2, max_col=7, max_row=ws_export.max_row):
            for cell in row:
                cell.value = None

        wb_analyse.save(fichier_dest)
        wb_analyse.close()

        print("🧹 Feuille 'export_timesheet' vidée (colonnes A à G, depuis ligne 2).")

        # 📥 Charger le fichier source
        wb_source = load_workbook(fichier_source)
        ws_source = wb_source["Sheet1"]

        # 📤 Recharger la feuille destination (déjà vidée)
        wb_dest = load_workbook(fichier_dest)
        ws_dest = wb_dest["export_timesheet"]

        # 📝 Copier lignes à partir de la ligne 2, colonnes A à G (1 à 7)
        row_dest = 2
        for row in ws_source.iter_rows(min_row=2, max_col=7):
            for col_index, cell in enumerate(row, start=1):
                ws_dest.cell(row=row_dest, column=col_index).value = cell.value
            row_dest += 1

        # 💾 Enregistrer et fermer
        wb_dest.save(fichier_dest)
        for wb in [wb_dest, wb_source, wb_analyse]:
            try:
                wb.close()
            except:
                pass

        print(f"✅ Données copiées dans 'export_timesheet' jusqu'à la ligne {row_dest - 1}")

        # 🕒 Obtenir date et heure d'exécution
        now = datetime.now()
        nouvelle_date = now.strftime("%Y%m%d")  # exemple : 20250726
        nouvelle_heure = arrondir_heure_par_15min(now)  # exemple : 1600

        ancien_nom = os.path.basename(fichier_dest)

        # 🔁 Remplacer la partie "extract 2025xxxx_xxxx" par la date et heure actuelles
        nouveau_nom = re.sub(
            r'extract \d{8}_\d{4}',
            f'extract {nouvelle_date}_{nouvelle_heure}',
            ancien_nom
        )
        chemin_nouveau = os.path.join(dossier_extract, nouveau_nom)

        # 📥 Écriture via fichier temporaire + remplacement atomique
        with tempfile.NamedTemporaryFile(delete=False, dir=dossier_extract, suffix=".xlsx") as tmp:
            chemin_tmp = tmp.name
        shutil.copy2(fichier_dest, chemin_tmp)
        os.replace(chemin_tmp, chemin_nouveau)

        log(f"✅ Fichier renommé proprement : {nouveau_nom}")

        sous_dossier = f"{annee[-2:]}-{mois}"
        dossier_sharepoint = os.path.join(
            os.path.expanduser("~"),
            "Wavestone",
            "WO - CTO - CDM - Clôture",
            sous_dossier
        )
        os.makedirs(dossier_sharepoint, exist_ok=True)

        chemin_final = os.path.join(dossier_sharepoint, nouveau_nom)
        shutil.copy2(chemin_nouveau, chemin_final)
        print(f"📤 Fichier copié dans : {chemin_final}")

        supprimer_et_attendre(chemin_nouveau)

        # 🧼 Nettoyage du dossier extract
        for fichier_temp in os.listdir(dossier_extract):
            chemin_temp = os.path.join(dossier_extract, fichier_temp)
            if os.path.isfile(chemin_temp):
                try:
                    os.remove(chemin_temp)
                    print(f"🗑️ Fichier supprimé dans 'extract' : {fichier_temp}")
                except Exception as e:
                    print(f"⚠️ Impossible de supprimer {fichier_temp} : {e}")

        print("✅ Dossier 'extract' vidé.")



    if choix == "Suivi du TACE Overrun":
        
        fichiers = glob.glob(os.path.join(dossier_extract, "*.xlsx"))
        fichiers_trie = sorted(fichiers, key=os.path.getctime, reverse=True)

        if len(fichiers_trie) < 2:
            print("❌ Moins de 2 fichiers trouvés pour la fusion.")
            return

        fichier_1, fichier_2 = fichiers_trie[:2]

        log(f"✅ Fusion de :\n- {os.path.basename(fichier_1)}\n- {os.path.basename(fichier_2)}")

        if "analyse des imputations" in os.path.basename(fichier_1).lower():
            fichier_dest = fichier_1
            fichier_source = fichier_2
        else:
            fichier_dest = fichier_2
            fichier_source = fichier_1

        # 🔄 Vider les colonnes A à R (1 à 18) de la feuille "export_overrun" à partir de la ligne 2
        wb_analyse = load_workbook(fichier_dest)
        if "export_overrun" not in wb_analyse.sheetnames:
            print("❌ Feuille 'export_overrun' introuvable.")
            return
        ws_export = wb_analyse["export_overrun"]

        for row in ws_export.iter_rows(min_row=2, max_col=18, max_row=ws_export.max_row):
            for cell in row:
                cell.value = None

        wb_analyse.save(fichier_dest)
        wb_analyse.close()
        print("🧹 Feuille 'export_overrun' vidée (colonnes A à R, depuis ligne 2).")

        # 📥 Charger le fichier source
        wb_source = load_workbook(fichier_source)
        ws_source = wb_source["Sheet1"]

        # 📤 Recharger la feuille destination (déjà vidée)
        wb_dest = load_workbook(fichier_dest)
        ws_dest = wb_dest["export_overrun"]

        row_dest = 2
        for row in ws_source.iter_rows(min_row=2, max_col=18):
            for col_index, cell in enumerate(row, start=1):
                ws_dest.cell(row=row_dest, column=col_index).value = cell.value
            row_dest += 1

        wb_dest.save(fichier_dest)

        # ✅ Fermer les fichiers correctement
        for wb in [wb_dest, wb_source, wb_analyse]:
            try:
                wb.close()
            except:
                pass

        print(f"✅ Données copiées dans 'export_overrun' jusqu'à la ligne {row_dest - 1}")

        # 🕒 Obtenir date et heure d'exécution
        now = datetime.now()
        nouvelle_date = now.strftime("%Y%m%d")  # exemple : 20250726
        nouvelle_heure = arrondir_heure_par_15min(now)  # exemple : 1600

        ancien_nom = os.path.basename(fichier_dest)

        # 🔁 Remplacer la partie "extract 2025xxxx_xxxx" par la date et heure actuelles
        nouveau_nom = re.sub(
            r'extract \d{8}_\d{4}',
            f'extract {nouvelle_date}_{nouvelle_heure}',
            ancien_nom
        )
        chemin_nouveau = os.path.join(dossier_extract, nouveau_nom)

        # 📥 Écriture via fichier temporaire puis remplacement pour éviter verrouillage
        with tempfile.NamedTemporaryFile(delete=False, dir=dossier_extract, suffix=".xlsx") as tmp:
            chemin_tmp = tmp.name
        shutil.copy2(fichier_dest, chemin_tmp)
        os.replace(chemin_tmp, chemin_nouveau)

        log(f"✅ Fichier renommé proprement : {nouveau_nom}")

        # 📁 Copie vers le bon dossier SharePoint
        sous_dossier = f"{annee[-2:]}-{mois}"
        dossier_sharepoint = os.path.join(
            os.path.expanduser("~"),
            "Wavestone",
            "WO - CTO - CDM - Clôture",
            sous_dossier
        )
        os.makedirs(dossier_sharepoint, exist_ok=True)
        chemin_final = os.path.join(dossier_sharepoint, nouveau_nom)
        shutil.copy2(chemin_nouveau, chemin_final)
        print(f"📤 Fichier copié dans : {chemin_final}")

        supprimer_et_attendre(chemin_nouveau)

        # 🧼 Nettoyage du dossier extract
        for fichier_temp in os.listdir(dossier_extract):
            chemin_temp = os.path.join(dossier_extract, fichier_temp)
            if os.path.isfile(chemin_temp):
                try:
                    os.remove(chemin_temp)
                    print(f"🗑️ Fichier supprimé dans 'extract' : {fichier_temp}")
                except Exception as e:
                    print(f"⚠️ Impossible de supprimer {fichier_temp} : {e}")

        print("✅ Dossier 'extract' vidé.")



    if choix == "Suivi des imputations non soumises":

        fichiers = glob.glob(os.path.join(dossier_extract, "*.xlsx"))
        fichiers_trie = sorted(fichiers, key=os.path.getctime, reverse=True)

        if len(fichiers_trie) < 2:
            print("❌ Moins de 2 fichiers trouvés pour la fusion.")
            return

        fichier_1, fichier_2 = fichiers_trie[:2]
        log(f"✅ Fusion de :\n- {os.path.basename(fichier_1)}\n- {os.path.basename(fichier_2)}")

        if "Suivi des imputations" in fichier_1:
            fichier_dest = fichier_1
            fichier_source = fichier_2
        else:
            fichier_dest = fichier_2
            fichier_source = fichier_1

        wb_dest = load_workbook(fichier_dest)
        ws_dest = wb_dest["Imputations non soumises"]

        wb_src = load_workbook(fichier_source)
        ws_src = wb_src["Sheet1"]

        row_dest = 2
        for row in ws_src.iter_rows(min_row=2, max_col=7):
            for col_index, cell in enumerate(row, start=1):
                ws_dest.cell(row=row_dest, column=col_index).value = cell.value
            row_dest += 1

        now = datetime.now()
        nouvelle_date = now.strftime("%Y%m%d")
        nouvelle_heure = arrondir_heure_par_15min(now)
        ancien_nom = os.path.basename(fichier_dest)

        nouveau_nom = re.sub(
            r'extract \d{8}_\d{4}',
            f'extract {nouvelle_date}_{nouvelle_heure}',
            ancien_nom
        )

        chemin_nouveau = os.path.join(dossier_extract, nouveau_nom)

        wb_dest.save(chemin_nouveau)
        wb_dest.close()
        wb_src.close()

        print(f"Ligne insérée : {row_dest - 1}")
        print(f"✅ Fusion terminée dans le fichier : {nouveau_nom}")
        print("✅ Fichier généré avec succès.")

        # 📤 Copier vers le sous-dossier de clôture
        sous_dossier = f"{annee[-2:]}-{mois}"
        dossier_sharepoint = os.path.join(
            os.path.expanduser("~"),
            "Wavestone",
            "WO - CTO - CDM - Clôture",
            sous_dossier,
            "Fichiers suivi"
        )
        os.makedirs(dossier_sharepoint, exist_ok=True)

        chemin_final = os.path.join(dossier_sharepoint, nouveau_nom)
        shutil.copy2(chemin_nouveau, chemin_final)
        print(f"📤 Copie réussie dans SharePoint :\n{chemin_final}")

        supprimer_et_attendre(chemin_nouveau)

        # 🧹 Supprimer les anciens fichiers "Suivi des imputations non soumises" dans le sous-dossier (sauf le nouveau)
        for fichier in os.listdir(dossier_sharepoint):
            chemin_fichier = os.path.join(dossier_sharepoint, fichier)
            if (
                fichier.startswith(f"{annee}{mois}- Suivi des imputations non soumises")
                and fichier != nouveau_nom
                and os.path.isfile(chemin_fichier)
            ):
                os.remove(chemin_fichier)
                print(f"🗑️ Ancien fichier supprimé : {fichier}")

        # 🧼 Vider le dossier "extract"
        gc.collect()
        time.sleep(1)
        for fichier_temp in os.listdir(dossier_extract): 
            chemin_temp = os.path.join(dossier_extract, fichier_temp)
            if os.path.isfile(chemin_temp):
                try:
                    os.remove(chemin_temp)
                    print(f"🗑️ Fichier supprimé dans 'extract' : {fichier_temp}")
                except Exception as e:
                    print(f"⚠️ Impossible de supprimer {fichier_temp} : {e}")
                    

         # 🧹 Vider également le dossier extract du SharePoint (hors sous-dossier)
        dossier_extract_sharepoint = os.path.join(
            os.path.expanduser("~"),
            "Wavestone",
            "WO - CTO - CDM - Clôture",
            "extract"
        )
        for fichier in os.listdir(dossier_extract_sharepoint):
            chemin_fichier = os.path.join(dossier_extract_sharepoint, fichier)
            if os.path.isfile(chemin_fichier):
                os.remove(chemin_fichier)
                print(f"🗑️ Fichier supprimé dans SharePoint/extract : {fichier}")

