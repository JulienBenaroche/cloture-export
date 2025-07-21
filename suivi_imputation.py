import os
import glob
from datetime import datetime
import shutil
import openpyxl

def executer(mois, annee):
    print()
    print("🔧 Lancement du module : Suivi des imputations non soumises")
    print(f"📅 Mois choisi : {mois}/{annee}")
    print()

    base_path = os.path.expanduser("~")
    sous_dossier = f"{annee[-2:]}-{mois}"
    dossier = os.path.join(
        base_path,
        "Wavestone",
        "WO - CTO - CDM - Clôture",
        sous_dossier,
        "Fichiers suivi"
    )

    print(f"📁 Chemin de recherche :\n{dossier}\n")

    if not os.path.exists(dossier):
        print(f"❌ Dossier introuvable : {dossier}")
        return

    pattern = os.path.join(
        dossier,
        f"{annee}{mois}- Suivi des imputations non soumises - extract *.xlsx"
    )

    print(f"🔍 Recherche de fichiers avec le motif :\n{pattern}\n")

    fichiers = glob.glob(pattern)

    if not fichiers:
        print("❌ Aucun fichier trouvé avec ce motif.")
        return

    def extraire_datetime(fichier):
        try:
            nom = os.path.basename(fichier)
            date_part = nom.split("extract ")[1].split("_")[0]
            heure_part = nom.split("extract ")[1].split("_")[1]
            return datetime.strptime(date_part + heure_part, "%Y%m%d%H%M")
        except Exception:
            print(f"⛔ Ignoré (format invalide) : {nom}")
            return datetime.min

    fichiers_trie = sorted(fichiers, key=extraire_datetime, reverse=True)
    fichier_plus_recent = fichiers_trie[0]

    print(f"✅ Fichier le plus récent trouvé :\n{os.path.basename(fichier_plus_recent)}")
    print(f"📍 Chemin complet :\n{fichier_plus_recent}\n")

    # 🔄 Dossier EXTRACT centralisé (SharePoint)
    dossier_extract = os.path.join(
        base_path,
        "Wavestone",
        "WO - CTO - CDM - Clôture",
        "extract"
    )
    os.makedirs(dossier_extract, exist_ok=True)

    nom_fichier = os.path.basename(fichier_plus_recent)
    chemin_copie = os.path.join(dossier_extract, nom_fichier)
    shutil.copy2(fichier_plus_recent, chemin_copie)

    print(f"📥 Copie locale enregistrée dans :\n{chemin_copie}")

    # ✅ Nettoyage sans toucher à la colonne H
    try:
        wb = openpyxl.load_workbook(chemin_copie)
        if "Imputations non soumises" in wb.sheetnames:
            ws = wb["Imputations non soumises"]

            for row in range(1, ws.max_row + 1):
                if ws.row_dimensions[row].hidden:
                    ws.row_dimensions[row].hidden = False

            # Vider colonnes A à G (1 à 7), ligne 2 et suivantes
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=7):
                for cell in row:
                    cell.value = None

            wb.save(chemin_copie)
            print("🧹 Feuille 'Imputations non soumises' nettoyée (colonnes A à G vidées, colonne H préservée).")
        else:
            print("⚠️ Feuille 'Imputations non soumises' introuvable.")
        wb.close()
    except Exception as e:
        print(f"❌ Erreur lors du traitement Excel : {e}")

    print("\n✅ Terminé.\n")
    return chemin_copie
